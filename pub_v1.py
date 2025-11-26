import pandas as pd
import numpy as np
from datetime import datetime
import re
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class DataLoader:
    """Handles loading and initial processing of CSV files."""
    
    def __init__(self, client_file_path, home_file_path):
        self.client_file_path = client_file_path
        self.home_file_path = home_file_path
        self.client_df = None
        self.home_df = None
    
    def load_files(self):
        """Load CSV files into pandas DataFrames."""
        try:
            self.client_df = pd.read_csv(self.client_file_path, encoding='utf-8')
            self.home_df = pd.read_csv(self.home_file_path, encoding='utf-8')
            
            # Strip whitespace from column names
            self.client_df.columns = self.client_df.columns.str.strip()
            self.home_df.columns = self.home_df.columns.str.strip()
            
            print(f"Client file loaded: {len(self.client_df)} rows")
            print(f"Home file loaded: {len(self.home_df)} rows")
            
            return self.client_df, self.home_df
            
        except FileNotFoundError as e:
            print(f"Error: File not found - {e}")
            sys.exit(1)
        except Exception as e:
            print(f"Error loading files: {e}")
            sys.exit(1)

class ExcelFormatter:
    """Handles Excel file formatting including cell colors."""
    
    def __init__(self, output_file_path):
        self.output_file_path = output_file_path
    
    def apply_colors(self):
        """
        Apply colors to cells:
        - Red: 'Not found'
        - Yellow: Revision mismatches (contains '/') and 'No Revision Date is given' in Note
        """
        try:
            wb = load_workbook(self.output_file_path)
            ws = wb.active
            
            # Define colors
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            
            # Find column indices
            headers = [cell.value for cell in ws[1]]
            
            try:
                result_col_idx = headers.index('Result') + 1
            except ValueError:
                result_col_idx = None
            
            try:
                note_col_idx = headers.index('Note') + 1
            except ValueError:
                note_col_idx = None
            
            # Apply colors
            for row_idx in range(2, ws.max_row + 1):
                if result_col_idx:
                    result_cell = ws.cell(row=row_idx, column=result_col_idx)
                    result_value = str(result_cell.value) if result_cell.value else ''
                    
                    # Red for 'Not found'
                    if 'Not found' in result_value:
                        result_cell.fill = red_fill
                    
                    # Yellow for mismatches
                    elif '/' in result_value and result_value != 'Verified':
                        result_cell.fill = yellow_fill
                
                # Yellow for Note column with 'No Revision Date is given'
                if note_col_idx:
                    note_cell = ws.cell(row=row_idx, column=note_col_idx)
                    note_value = str(note_cell.value) if note_cell.value else ''
                    
                    if 'No Revision Date is given' in note_value:
                        note_cell.fill = yellow_fill
            
            wb.save(self.output_file_path)
            wb.close()
            print("Cell colors applied successfully!")
            
        except Exception as e:
            print(f"Warning: Could not apply cell colors: {e}")

class ClientFormatter:
    """Handles formatting of client file based on business rules."""
    
    def __init__(self, client_df):
        self.client_df = client_df.copy()
    
    def clean_revision_no(self):
        """Clean Revision No. column by:
        1. Remove characters after first comma, EXCEPT if followed by 'STATEMENT'
        2. Remove leading zeros from numeric parts
        
        Examples:
        '03, TR 005, -006' -> '3'
        '5, TR01' -> '5'
        '3, STATEMENT 5214' -> '3, STATEMENT 5214'
        '25, TR 25-16' -> '25'
        '02' -> '2'
        """
        print("===== THIS IS CLEANING REVISION NO COLUMN =====")
        
        if 'Revision No.' not in self.client_df.columns:
            print("Warning: 'Revision No.' column not found in client file")
            return self
        
        def process_single_value(rev_str):
            if pd.isna(rev_str) or str(rev_str).strip() == '':
                return ''
            
            rev_str = str(rev_str).strip()
            original = rev_str
            
            # Step 1: Remove characters after first comma, except if followed by 'STATEMENT'
            if ',' in rev_str:
                parts = rev_str.split(',', 1)  # Split only on first comma
                first_part = parts[0].strip()
                second_part = parts[1].strip() if len(parts) > 1 else ''
                
                # Check if second part starts with STATEMENT (case insensitive)
                if second_part.upper().startswith('STATEMENT'):
                    rev_str = f"{first_part}, {second_part}"
                else:
                    rev_str = first_part
            
            # Step 2: Remove leading zeros from numeric parts
            # Handle cases with STATEMENT separately
            if 'STATEMENT' in rev_str.upper():
                # Split by comma to process the numeric part
                parts = rev_str.split(',', 1)
                if len(parts) >= 1:
                    numeric_part = parts[0].strip()
                    # Remove leading zeros
                    numeric_part = numeric_part.lstrip('0') or '0'
                    # Reconstruct with STATEMENT part
                    rev_str = f"{numeric_part}, {parts[1].strip()}"
            else:
                # Simple case: just remove leading zeros
                rev_str = rev_str.lstrip('0') or '0'
            
            print(f"  Cleaned: '{original}' -> '{rev_str}'")
            return rev_str
        
        # Apply cleaning to the entire column
        print("\nCleaning Revision No. column...")
        for idx in self.client_df.index:
            original = self.client_df.at[idx, 'Revision No.']
            cleaned = process_single_value(original)
            self.client_df.at[idx, 'Revision No.'] = cleaned
        
        print("===== REVISION NO. CLEANING COMPLETE =====\n")
        return self
    
    def create_formatted_column(self):
        """
        Create 'Formatted' column based on 'Revision No.' rules:
        1. If contains TR or '-', set to 'TR'
        2. If contains STATEMENT, extract the number after it
        3. Ignore rows with more than 1 comma
        4. If contains 'TR' after comma, extract TR value
        """
        print("===== THIS IS CREATING FORMAT COLUMN =====")
        
        self.client_df['Formatted'] = ''
        
        for idx, row in self.client_df.iterrows():
            rev_no = row.get('Revision No.')
            
            if pd.isna(rev_no):
                continue
            
            rev_str = str(rev_no).strip()
            
            # Count commas
            comma_count = rev_str.count(',')
            
            # Rule 3: Ignore if more than 1 comma
            if comma_count > 1:
                continue
            
            # Rule 2: Handle STATEMENT
            if 'STATEMENT' in rev_str.upper():
                match = re.search(r'STATEMENT\s+([\d\-A-Z]+)', rev_str, re.IGNORECASE)
                if match:
                    self.client_df.at[idx, 'Formatted'] = match.group(1)
                continue
            
            # Rule 4: Handle TR after comma
            if comma_count == 1 and 'TR' in rev_str.upper():
                # Extract everything after comma
                parts = rev_str.split(',')
                if len(parts) == 2:
                    after_comma = parts[1].strip()
                    # Extract TR and following pattern
                    tr_match = re.search(r'(TR\s*[\d\-]+)', after_comma, re.IGNORECASE)
                    if tr_match:
                        self.client_df.at[idx, 'Formatted'] = tr_match.group(1).strip()
                        continue
            
            # Rule 1: If contains TR or '-', set to 'TR'
            if 'TR' in rev_str.upper() or '-' in rev_str:
                self.client_df.at[idx, 'Formatted'] = 'TR'
        
        print("===== FORMAT COLUMN CREATION COMPLETE =====\n")
        return self.client_df
    
    def process(self):
        """Process client file: preprocess columns first, then create formatted column."""
        print("\n" + "="*60)
        print("STARTING CLIENT FILE PROCESSING")
        print("="*60)
        
        # Step 1: Clean the Revision No. column
        self.clean_revision_no()
        
        # Step 2: Create the Formatted column
        self.create_formatted_column()
        
        print("="*60)
        print("CLIENT FILE PROCESSING COMPLETE")
        print("="*60 + "\n")
        
        return self.client_df
    
    def save_formatted_file(self, output_path='client_formatted.csv'):
        """Save the formatted client file."""
        try:
            self.client_df.to_csv(output_path, index=False)
            print(f"Formatted client file saved to: {output_path}")
        except Exception as e:
            print(f"Error saving formatted file: {e}")

class HomeProcessor:
    """Handles home file processing including duplicate removal."""
    
    def __init__(self, home_df):
        self.home_df = home_df.copy()
    
    @staticmethod
    def remove_leading_zeros(value_str):
        """Remove leading zeros from a string value.
        
        Examples:
        '02' -> '2'
        '04' -> '4'
        '70' -> '70'
        '0' -> '0'
        """
        if pd.isna(value_str) or str(value_str).strip() == '':
            return ''
        
        value_str = str(value_str).strip()
        
        # Remove leading zeros, but keep at least one digit
        result = value_str.lstrip('0') or '0'
        
        return result
    
    def preprocess_revision_num(self):
        """Preprocess Revision Num column by removing leading zeros."""
        if 'Revision Num' not in self.home_df.columns:
            print("Warning: 'Revision Num' column not found in home file")
            return self
        
        print("\n=== Preprocessing Home Revision Num ===")
        
        for idx in self.home_df.index:
            original = self.home_df.at[idx, 'Revision Num']
            cleaned = self.remove_leading_zeros(original)
            if str(original) != str(cleaned):
                print(f"  Row {idx}: '{original}' -> '{cleaned}'")
            self.home_df.at[idx, 'Revision Num'] = cleaned
        
        print("=== Home Revision Num Preprocessing Complete ===\n")
        return self
    
    def remove_duplicates(self):
        """
        Remove duplicate rows based on 'Call Number' and 'Revision Description'.
        Keeps the first occurrence.
        """
        print("\n" + "="*70)
        print("DUPLICATE REMOVAL FROM HOME FILE")
        print("="*70)
        
        required_columns = ['Call Number', 'Revision Description']
        missing_columns = [col for col in required_columns if col not in self.home_df.columns]
        
        if missing_columns:
            print(f"Warning: Missing columns: {missing_columns}")
            print("Skipping duplicate removal.")
            print("="*70 + "\n")
            return self.home_df
        
        original_count = len(self.home_df)
        
        # Create index for tracking
        self.home_df['original_index'] = range(len(self.home_df))
        
        # Identify duplicates
        duplicates_mask = self.home_df.duplicated(
            subset=['Call Number', 'Revision Description'], 
            keep='first'
        )
        
        duplicate_rows = self.home_df[duplicates_mask].copy()
        
        if len(duplicate_rows) > 0:
            self.home_df = self.home_df[~duplicates_mask].copy()
            self.home_df = self.home_df.drop(columns=['original_index'])
            self.home_df = self.home_df.reset_index(drop=True)
            
            print(f"Summary: Removed {original_count - len(self.home_df)} duplicate(s)")
            print(f"Home file now contains: {len(self.home_df)} rows")
        else:
            print("No duplicates found")
            print(f"Home file contains: {len(self.home_df)} rows")
            self.home_df = self.home_df.drop(columns=['original_index'])
        
        print("="*70 + "\n")
        return self.home_df
    
    def process(self):
        """Process home file: preprocess columns first, then remove duplicates."""
        self.preprocess_revision_num()
        self.remove_duplicates()
        return self.home_df


# class ClientFormatter:
#     """Handles formatting of client file based on business rules."""
    
#     def __init__(self, client_df):
#         self.client_df = client_df.copy()
    
#     def create_formatted_column(self):
#         """
#         Create 'Formatted' column based on 'Revision No.' rules:
#         1. If contains TR or '-', set to 'TR'
#         2. If contains STATEMENT, extract the number after it
#         3. Ignore rows with more than 1 comma
#         4. If contains 'TR' after comma, extract TR value
#         """
#         self.client_df['Formatted'] = ''
        
#         for idx, row in self.client_df.iterrows():
#             rev_no = row.get('Revision No.')
            
#             if pd.isna(rev_no):
#                 continue
            
#             rev_str = str(rev_no).strip()
            
#             # Count commas
#             comma_count = rev_str.count(',')
            
#             # Rule 3: Ignore if more than 1 comma
#             if comma_count > 1:
#                 continue
            
#             # Rule 2: Handle STATEMENT
#             if 'STATEMENT' in rev_str.upper():
#                 match = re.search(r'STATEMENT\s+([\d\-A-Z]+)', rev_str, re.IGNORECASE)
#                 if match:
#                     self.client_df.at[idx, 'Formatted'] = match.group(1)
#                 continue
            
#             # Rule 4: Handle TR after comma
#             if comma_count == 1 and 'TR' in rev_str.upper():
#                 # Extract everything after comma
#                 parts = rev_str.split(',')
#                 if len(parts) == 2:
#                     after_comma = parts[1].strip()
#                     # Extract TR and following pattern
#                     tr_match = re.search(r'(TR\s*[\d\-]+)', after_comma, re.IGNORECASE)
#                     if tr_match:
#                         self.client_df.at[idx, 'Formatted'] = tr_match.group(1).strip()
#                         continue
            
#             # Rule 1: If contains TR or '-', set to 'TR'
#             if 'TR' in rev_str.upper() or '-' in rev_str:
#                 self.client_df.at[idx, 'Formatted'] = 'TR'
        
#         return self.client_df
    
#     def save_formatted_file(self, output_path='client_formatted.csv'):
#         """Save the formatted client file."""
#         try:
#             self.client_df.to_csv(output_path, index=False)
#             print(f"Formatted client file saved to: {output_path}")
#         except Exception as e:
#             print(f"Error saving formatted file: {e}")


# class HomeProcessor:
#     """Handles home file processing including duplicate removal."""
    
#     def __init__(self, home_df):
#         self.home_df = home_df.copy()
    
#     def remove_duplicates(self):
#         """
#         Remove duplicate rows based on 'Call Number' and 'Revision Description'.
#         Keeps the first occurrence.
#         """
#         print("\n" + "="*70)
#         print("DUPLICATE REMOVAL FROM HOME FILE")
#         print("="*70)
        
#         required_columns = ['Call Number', 'Revision Description']
#         missing_columns = [col for col in required_columns if col not in self.home_df.columns]
        
#         if missing_columns:
#             print(f"Warning: Missing columns: {missing_columns}")
#             print("Skipping duplicate removal.")
#             print("="*70 + "\n")
#             return self.home_df
        
#         original_count = len(self.home_df)
        
#         # Create index for tracking
#         self.home_df['original_index'] = range(len(self.home_df))
        
#         # Identify duplicates
#         duplicates_mask = self.home_df.duplicated(
#             subset=['Call Number', 'Revision Description'], 
#             keep='first'
#         )
        
#         duplicate_rows = self.home_df[duplicates_mask].copy()
        
#         if len(duplicate_rows) > 0:
#             # print(f"Found {len(duplicate_rows)} duplicate row(s) to remove:\n")
            
#             for idx, row in duplicate_rows.iterrows():
#                 original_idx = row['original_index']
#                 # print(f"Row {original_idx + 2} (Excel row):")
#                 # print(f"  - Call Number: {row['Call Number']}")
#                 # print(f"  - Revision Description: {row['Revision Description']}")
#                 # print(f"  - Document Number: {row.get('Document Number', 'N/A')}")
#                 # print()
            
#             self.home_df = self.home_df[~duplicates_mask].copy()
#             self.home_df = self.home_df.drop(columns=['original_index'])
#             self.home_df = self.home_df.reset_index(drop=True)
            
#             print(f"Summary: Removed {original_count - len(self.home_df)} duplicate(s)")
#             print(f"Home file now contains: {len(self.home_df)} rows")
#         else:
#             print("No duplicates found")
#             print(f"Home file contains: {len(self.home_df)} rows")
#             self.home_df = self.home_df.drop(columns=['original_index'])
        
#         print("="*70 + "\n")
#         return self.home_df



